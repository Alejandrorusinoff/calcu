#Aplicacion de escritorio ------ Tabla de pedidos------

#--------bibliotecas--------
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk
from datetime import date
from datetime import datetime



#--------funciones--------

def validar(a):
    if a == "":
        a = a.replace("","0")
        a = int(a)
    else:
        a = int(a)
    return a

def name_pedido():
        nombr_pedido = tk.Tk()
        nombr_pedido.title("Nombre del pedido")
        nombr_pedido.config(width=200, height=110)

        nombre_pedido=ttk.Label(nombr_pedido, text="ingrese el nombre del pedido")
        nombre_pedido.place(x=10, y=10)

        c_nombre_pedido=ttk.Entry(nombr_pedido)
        c_nombre_pedido.place(x=10, y=35, width=180, height=23)

        bot_guardar=ttk.Button(nombr_pedido, text= "Guardar")
        bot_guardar.place(x=60, y=70)

        nombre__pedido = c_nombre_pedido.get()
        fecha = date.today()
        #hora = time.strftime("%H:%M:%S")
        nombre_archivo = nombre__pedido + "pedidos_" + str(fecha)
        if nombre__pedido !="":
            return nombre_archivo
        



def añadir_a_base_de_datos():
    xc_autoperforantes = validar(c_autoperforantes.get())
    xc_borneras2_5 = validar(c_borneras2_5.get())
    xc_borneras4 = validar(c_borneras4.get())
    xc_borneras6 = validar(c_borneras6.get())
    xc_cable0_75 = validar(c_cable0_75.get())
    xc_cable2_5 = validar(c_cable2_5.get())
    xc_cable6 = validar(c_cable6.get())
    xc_cable_canal_40x70 = validar(c_cable_canal_40x70.get())
    xc_cable_canal_70x70 = validar(c_cable_canal_70x70.get())
    xc_fuente12 = validar(c_fuente12.get())
    xc_fuente24 = validar(c_fuente24.get())
    xc_identificadores = validar(c_identificadores.get())
    xc_rele_doble_12vcc = validar(c_rele_doble_12vcc.get())
    xc_rele_doble_24vcc = validar(c_rele_doble_24vcc.get())
    xc_rele_cuadru_12vcc = validar(c_rele_cuadru_12vcc.get())
    xc_rele_cuadru_24vcc = validar(c_rele_cuadru_24vcc.get())
    xc_rieldin = validar(c_rieldin.get())
    xc_terminales_sin_leng_0_75 = validar(c_terminales_sin_leng_0_75.get())
    xc_terminales_con_leng_0_75 = validar(c_terminales_con_leng_0_75.get())
    xc_terminales_con_leng_2_5 = validar(c_terminales_con_leng_2_5.get())
    xc_zocaro_rele_doble = validar(c_zocaro_rele_doble.get())
    xc_zocaro_rele_cuadru = validar(c_zocaro_rele_cuadru.get())
    lista_materiales = [(2,xc_autoperforantes), (3,xc_borneras2_5), (4,xc_borneras4), (5,xc_borneras6), (6,xc_cable0_75),(7,xc_cable2_5), (8,xc_cable6), (9,xc_cable_canal_40x70), (10,xc_cable_canal_70x70), (11,xc_fuente12), (12,xc_fuente24), (13,xc_identificadores), (14,xc_rele_doble_12vcc), (15,xc_rele_doble_24vcc), (16,xc_rele_cuadru_12vcc), (17,xc_rele_cuadru_24vcc), (18,xc_rieldin), (19,xc_terminales_sin_leng_0_75), (20,xc_terminales_con_leng_0_75), (21,xc_terminales_con_leng_2_5), (22,xc_zocaro_rele_doble), (23,xc_zocaro_rele_cuadru)]
    
    # lista de materiales añadido para que se suma a la base de datos
    lista_materiales1 = [xc_autoperforantes,xc_borneras2_5, xc_borneras4, xc_borneras6, xc_cable0_75, xc_cable2_5, xc_cable6, xc_cable_canal_40x70, xc_cable_canal_70x70, xc_fuente12, xc_fuente24, xc_identificadores, xc_rele_doble_12vcc, xc_rele_doble_24vcc, xc_rele_cuadru_12vcc, xc_rele_cuadru_24vcc, xc_rieldin, xc_terminales_sin_leng_0_75, xc_terminales_con_leng_0_75, xc_terminales_con_leng_2_5, xc_zocaro_rele_doble, xc_zocaro_rele_cuadru]
    
    #abro el excel leo la cantidad de materiales que hay en la hoja stock de materia prima
    wb = load_workbook(filename = "C:\\Users\\User\\Desktop\\Ale\\Devops\\Python promaming\\Clase 7- Desarrollo de aplicaciones de escritorio\\Ejercicio_integrador\\Cant_mp.xlsx")
    ws = wb.active
    libro = wb.get_sheet_by_name("Stock_de_materia_prima")
    cant_mp = []
    for i in range(2,24):
        celdas=libro.cell(row = i, column = 2).value
        cant_mp.append(celdas)

    s = [i+j for i,j in zip(lista_materiales1,cant_mp)] #suma la lista (lista_materiales1 y cant_mp)
    print(type(s))
    print(s)
    #wb.close()


    enumeracion = []
    for ii in range(0,21):
        enumeracion.append(ii)
    print(enumeracion)

    c = []

    for i, j in enumerate(s):
        c.append([str(enumeracion[i])+str(s[i])])
    print(c) 
    

    #abro el excel y guardo los datos
    #wb = load_workbook(filename = "C:\\Users\\User\\Desktop\\Ale\\Devops\\Python promaming\\Clase 7- Desarrollo de aplicaciones de escritorio\\Ejercicio_integrador\\Cant_mp.xlsx")
    
    
    for fila, fila_mat in c:
        d = ws.cell(row=fila, column=2, value=fila_mat)

    wb.save(filename = "C:\\Users\\User\\Desktop\\Ale\\Devops\\Python promaming\\Clase 7- Desarrollo de aplicaciones de escritorio\\Ejercicio_integrador\\Cant_mp.xlsx")
    wb.close()


#borrar materiales de la base de datos
def borrar_mat_base_datos():
    ventana_borrar_materiales = tk.Tk()
    ventana_borrar_materiales.title("Borrar materiales")
    ventana_borrar_materiales.config(width=200, height=250)

    usuario=ttk.Label(ventana_borrar_materiales, text="ingrese su usuario:")
    usuario.place(x=10, y=10)

    c_usuario=ttk.Entry(ventana_borrar_materiales)
    c_usuario.place(x=10, y=35, width=180, height=23)

    contraseña=ttk.Label(ventana_borrar_materiales, text="ingrese su contraseña:")
    contraseña.place(x=10, y=80)

    c_contraseña=ttk.Entry(ventana_borrar_materiales)
    c_contraseña.place(x=10, y=105, width=180, height=23)

    #boton_aceptar=ttk.Button(ventana_borrar_materiales, text= "Aceptar", command=aceptar)
    #boton_aceptar.place(x=60, y=150)

    #obtener los valores de las 2 cajas de la ventana_borrar_materiales
    obt_usuario = c_usuario.get()
    obt_contraseña = c_contraseña.get()


    
    if obt_usuario == "alejandrorusinoff" and obt_contraseña == "36296791":
        wb = load_workbook(filename = "C:\\Users\\User\\Desktop\\Ale\\Devops\\Python promaming\\Clase 7- Desarrollo de aplicaciones de escritorio\\Ejercicio_integrador\\Cant_mp.xlsx")
        ws = wb.active
            
        for fila in range(2,24):
            d = ws.cell(row=fila, column=2, value="")

        wb.save(filename = "C:\\Users\\User\\Desktop\\Ale\\Devops\\Python promaming\\Clase 7- Desarrollo de aplicaciones de escritorio\\Ejercicio_integrador\\Cant_mp.xlsx")
        wb.close()

    elif obt_usuario != "" and obt_contraseña != "" and obt_contraseña != "alejandrorusinoff" and obt_contraseña != "36296791":
        error=ttk.Label(ventana_borrar_materiales, text="usuario o contraseña incorrecta")
        error.place(x=10, y=190)

    else:
        pass




    #analizar xq no anda
    """
    #validacion
    while True:


        if obt_usuario == "alejandrorusinoff" and obt_contraseña == "36296791":
            ejecuto = aceptar()
            break

        else:
            pass
            break
    """
              
    ventana_borrar_materiales.mainloop()



def realizar_pedidos():
    xc_autoperforantes = c_autoperforantes.get()
    xc_borneras2_5 = c_borneras2_5.get()
    xc_borneras4 = c_borneras4.get()
    xc_borneras6 = c_borneras6.get()
    xc_cable0_75 = c_cable0_75.get()
    xc_cable2_5 = c_cable2_5.get()
    xc_cable6 = c_cable6.get()
    xc_cable_canal_40x70 = c_cable_canal_40x70.get()
    xc_cable_canal_70x70 = c_cable_canal_70x70.get()
    xc_fuente12 = c_fuente12.get()
    xc_fuente24 = c_fuente24.get()
    xc_identificadores = c_identificadores.get()
    xc_rele_doble_12vcc = c_rele_doble_12vcc.get()
    xc_rele_doble_24vcc = c_rele_doble_24vcc.get()
    xc_rele_cuadru_12vcc = c_rele_cuadru_12vcc.get()
    xc_rele_cuadru_24vcc = c_rele_cuadru_24vcc.get()
    xc_rieldin = c_rieldin.get()
    xc_terminales_sin_leng_0_75 = c_terminales_sin_leng_0_75.get()
    xc_terminales_con_leng_0_75 = c_terminales_con_leng_0_75.get()
    xc_terminales_con_leng_2_5 = c_terminales_con_leng_2_5.get()
    xc_zocaro_rele_doble = c_zocaro_rele_doble.get()
    xc_zocaro_rele_cuadru = c_zocaro_rele_cuadru.get()



    wb = Workbook()
    ws = wb.active
    ws.title = "Planilla de pedidos" #nombre de la hoja
    ws.append(["Materiales","cantidad"])
    
    lista_mat_pedir = [(2,"autoperforantes T1"), (3,"borneras2_5"), (4,"borneras4"), (5,"borneras6"), (6,"cable0_75"),(7,"cable2_5"), (8,"cable6"), (9,"cable_canal_40x70"), (10,"cable_canal_70x70"), (11,"fuente12"), (12,"fuente24"), (13,"identificadores"), (14,"rele_doble_12vcc"), (15,"rele_doble_24vcc"), (16,"rele_cuadru_12vcc"), (17,"rele_cuadru_24vcc"), (18,"rieldin"), (19,"terminales_sin_leng_0_75"), (20,"terminales_con_leng_0_75"), (21,"terminales_con_leng_2_5"), (22,"zocaro_rele_doble"), (23,"zocaro_rele_cuadru")]
    for idi, lista_mat in lista_mat_pedir:
        d = ws.cell(row=idi, column=1, value=lista_mat)

    cantidad_mat = [(2,xc_autoperforantes), (3,xc_borneras2_5), (4,xc_borneras4), (5,xc_borneras6), (6,xc_cable0_75),(7,xc_cable2_5), (8,xc_cable6), (9,xc_cable_canal_40x70), (10,xc_cable_canal_70x70), (11,xc_fuente12), (12,xc_fuente24), (13,xc_identificadores), (14,xc_rele_doble_12vcc), (15,xc_rele_doble_24vcc), (16,xc_rele_cuadru_12vcc), (17,xc_rele_cuadru_24vcc), (18,xc_rieldin), (19,xc_terminales_sin_leng_0_75), (20,xc_terminales_con_leng_0_75), (21,xc_terminales_con_leng_2_5), (22,xc_zocaro_rele_doble), (23,xc_zocaro_rele_cuadru)]
    for idi2, cant_mat in cantidad_mat:
        e = ws.cell(row=idi2, column=2, value=cant_mat)

    #crea caja de texto para el nombre
    nombre_del_pedido = name_pedido()  
    print(nombre_del_pedido)

    wb.save(filename="C:\\Users\\User\\Desktop\\Ale\\Devops\\Python promaming\\Clase 7- Desarrollo de aplicaciones de escritorio\\Ejercicio_integrador\\"+ nombre_del_pedido +".xlsx")
    wb.close()




def nuevo():
    print("Nuevo archivo")

def acerca_de():
    print("Acerca de")

ventana_principal = tk.Tk()
ventana_principal.title("Pedidos")


#Barras de menú
#1° paso crear barra de menus con Tk.Menu()
#2° paso crear manú con Tk.Menu(....)
#3° paso crear comandos de los menus con .add_command
#4° paso agregar los menus a la barra de menus con .add_cascade
#5° paso indicar que la barra de menú estará en la ventana con .config(...)

barra_de_menu = tk.Menu()
menu_archivo = tk.Menu(barra_de_menu, tearoff=0)
menu_archivo.add_command(label="Salir", command=ventana_principal.destroy)

menu_ayuda = tk.Menu(barra_de_menu, tearoff=0)
menu_ayuda.add_command(label="Acerca de...", command=acerca_de)

barra_de_menu.add_cascade(label="Archivo", menu=menu_archivo)
barra_de_menu.add_cascade(label="Ayuda", menu=menu_ayuda)

ventana_principal.config(width=900, height=900, menu=barra_de_menu)



#Lista desplegable
select_TTA = ttk.Combobox(
    values=["TTA 50KVA",
            "TTA 100KVA",
            "TTA 150KVA",
            "TTA 200KVA"
    ]
)

select_TTA.place(x=10, y=10)
select_TTA.current(0)


#Botones
boton_add_mp = ttk.Button(text= "Añadir materiales a la base de datos", command= añadir_a_base_de_datos)
boton_add_mp.place(x=500, y=10, width=200, height=25)

boton_add_mp = ttk.Button(text= "Borrar materiales de la base de datos", command= borrar_mat_base_datos)
boton_add_mp.place(x=500, y=50, width=200, height=25)

boton_stock = ttk.Button(text="Visualizar stock disponible")
boton_stock.place(x=500, y=90, width=200, height=25)

verif_pedidos = ttk.Button(text="Verificar si hay stock suficiente \n       para realizar el tablero")
verif_pedidos.place(x=500, y=130, width=200, height=50)

boton_pedidos = ttk.Button(text="Realizar pedidos", command=realizar_pedidos)
boton_pedidos.place(x=500, y=195, width=200, height=25)

#Etiquetas
autoperforantes = ttk.Label(text="Autoperforantes T1")
autoperforantes.place(x=10, y=50)

borneras2_5 = ttk.Label(text="Borneras 2,5mm")
borneras2_5.place(x=10, y=80)

borneras4 = ttk.Label(text="Borneras 4mm")
borneras4.place(x=10, y=110)

borneras6 = ttk.Label(text="Borneras 6mm")
borneras6.place(x=10, y=140)

cable0_75 = ttk.Label(text="Cable de 0,75mm")
cable0_75.place(x=10, y=170)

cable2_5 = ttk.Label(text="Cable de 2,5mm")
cable2_5.place(x=10, y=200)

cable6 = ttk.Label(text="Cables de 6mm")
cable6.place(x=10, y=230)

cable_canal_40x70 = ttk.Label(text="Cable canal ranurado de 40x70mm")
cable_canal_40x70.place(x=10, y=260)

cable_canal_70x70 = ttk.Label(text="Cable canal ranurado de 70x70mm")
cable_canal_70x70.place(x=10, y=290)

fuente12 = ttk.Label(text="Fuenta de alimentación 12Vcc 10A")
fuente12.place(x=10, y=320)

fuente24 = ttk.Label(text="Fuente de alimentación 24Vcc 10A")
fuente24.place(x=10, y=350)

identificadores = ttk.Label(text="Identificadores")
identificadores.place(x=10, y=380)

rele_doble_12vcc = ttk.Label(text="Rele doble inversor 12Vcc")
rele_doble_12vcc.place(x=10, y=410)

rele_doble_24vcc = ttk.Label(text="Rele doble inversor 24Vcc")
rele_doble_24vcc.place(x=10, y=440)

rele_cuadru_12vcc = ttk.Label(text="Rele cuadruple inversor 12Vcc")
rele_cuadru_12vcc.place(x=10, y=470)

rele_cuadru_24vcc = ttk.Label(text="Rele cuadruple inversor 24Vcc")
rele_cuadru_24vcc.place(x=10, y=470)

rieldin = ttk.Label(text="Rieldin")
rieldin.place(x=10, y=500)

terminales_sin_leng_0_75 = ttk.Label(text="Terminales punteras 0,75mm")
terminales_sin_leng_0_75.place(x=10, y=530)

terminales_con_leng_0_75 = ttk.Label(text="Terminales con lengueta 0,75mm")
terminales_con_leng_0_75.place(x=10, y=560)

terminales_con_leng_2_5 = ttk.Label(text="Terminales con lengueta 2,5mm")
terminales_con_leng_2_5.place(x=10, y=590) 

zocaro_rele_doble = ttk.Label(text="Zocalo de rele doble inversor")
zocaro_rele_doble.place(x=10, y=620)

zocaro_rele_cuadru = ttk.Label(text="Zocalo de rele cuadruple inversor")
zocaro_rele_cuadru.place(x=10, y=650)



#2da columna ----cantidad de materiales----

#etiqueta de cantidad de materiales
cant_mp = ttk.Label(text="Cantidad de Materiales")
cant_mp.place(x=260, y=10)

#caja de texto cantidad de materiales
c_autoperforantes = ttk.Entry()
c_autoperforantes.place(x=260, y=50)

c_borneras2_5 = ttk.Entry()
c_borneras2_5.place(x=260, y=80)

c_borneras4 = ttk.Entry()
c_borneras4.place(x=260, y=110)

c_borneras6 = ttk.Entry()
c_borneras6.place(x=260, y=140)

c_cable0_75 = ttk.Entry()
c_cable0_75.place(x=260, y=170)

c_cable2_5 = ttk.Entry()
c_cable2_5.place(x=260, y=200)

c_cable6 = ttk.Entry()
c_cable6.place(x=260, y=230)

c_cable_canal_40x70 = ttk.Entry()
c_cable_canal_40x70.place(x=260, y=260)

c_cable_canal_70x70 = ttk.Entry()
c_cable_canal_70x70.place(x=260, y=290)

c_fuente12 = ttk.Entry()
c_fuente12.place(x=260, y=320)

c_fuente24 = ttk.Entry()
c_fuente24.place(x=260, y=350)

c_identificadores = ttk.Entry()
c_identificadores.place(x=260, y=380)

c_rele_doble_12vcc = ttk.Entry()
c_rele_doble_12vcc.place(x=260, y=410)

c_rele_doble_24vcc = ttk.Entry()
c_rele_doble_24vcc.place(x=260, y=440)

c_rele_cuadru_12vcc = ttk.Entry()
c_rele_cuadru_12vcc.place(x=260, y=470)

c_rele_cuadru_24vcc = ttk.Entry()
c_rele_cuadru_24vcc.place(x=260, y=470)

c_rieldin = ttk.Entry()
c_rieldin.place(x=260, y=500)

c_terminales_sin_leng_0_75 = ttk.Entry()
c_terminales_sin_leng_0_75.place(x=260, y=530)

c_terminales_con_leng_0_75 = ttk.Entry()
c_terminales_con_leng_0_75.place(x=260, y=560)

c_terminales_con_leng_2_5 = ttk.Entry()
c_terminales_con_leng_2_5.place(x=260, y=590) 

c_zocaro_rele_doble = ttk.Entry()
c_zocaro_rele_doble.place(x=260, y=620)

c_zocaro_rele_cuadru = ttk.Entry()
c_zocaro_rele_cuadru.place(x=260, y=650)

#3er columnas ---- unidades----
#etuiqueta de unidades
unidades = ttk.Label(text="[U]")
unidades.place(x=423, y=10)

#listas desplegables
lista_despl_autoperforantes = ttk.Combobox(values=["Kg","U"])
lista_despl_autoperforantes.place(x=410, y=50, width=50, height=20)
lista_despl_autoperforantes.current(0)

lista_despl_borneras2_5 = ttk.Combobox(values=["Caja","U"])
lista_despl_borneras2_5.place(x=410, y=80, width=50, height=20 )
lista_despl_borneras2_5.current(0)

lista_despl_borneras4 = ttk.Combobox(values=["Caja","U"])
lista_despl_borneras4.place(x=410, y=110, width=50, height=20)
lista_despl_borneras4.current(0)

lista_despl_borneras6 = ttk.Combobox(values=["Caja","U"])
lista_despl_borneras6.place(x=410, y=140, width=50, height=20)
lista_despl_borneras6.current(0)

lista_despl_cable0_75 = ttk.Combobox(values=["Rollo","mts"])
lista_despl_cable0_75.place(x=410, y=170, width=50, height=20)
lista_despl_cable0_75.current(0)

lista_despl_cable2_5 = ttk.Combobox(values=["Rollo","mts"])
lista_despl_cable2_5.place(x=410, y=200, width=50, height=20)
lista_despl_cable2_5.current(0)

lista_despl_cable6 = ttk.Combobox(values=["Rollo","mts"])
lista_despl_cable6.place(x=410, y=230, width=50, height=20)
lista_despl_cable6.current(0)

lista_despl_cable_canal_40x70 = ttk.Combobox(values=["U","mts"])
lista_despl_cable_canal_40x70.place(x=410, y=260, width=50, height=20)
lista_despl_cable_canal_40x70.current(0)

lista_despl_cable_canal_70x70 = ttk.Combobox(values=["Caja","U"])
lista_despl_cable_canal_70x70.place(x=410, y=290, width=50, height=20)
lista_despl_cable_canal_70x70.current(0)

lista_despl_fuente12 = ttk.Label(text="U")
lista_despl_fuente12.place(x=410, y=320, width=50, height=20)

lista_despl_fuente24 = ttk.Label(text="U")
lista_despl_fuente24.place(x=410, y=350, width=50, height=20)

lista_despl_identificadores = ttk.Combobox(values=["Caja","U"])
lista_despl_identificadores.place(x=410, y=380, width=50, height=20)
lista_despl_identificadores.current(0)

lista_despl_rele_doble_inversor_12vcc = ttk.Combobox(values=["Caja","U"])
lista_despl_rele_doble_inversor_12vcc.place(x=410, y=410, width=50, height=20)
lista_despl_rele_doble_inversor_12vcc.current(0)

lista_despl_rele_doble_inversor_24vcc = ttk.Combobox(values=["Caja","U"])
lista_despl_rele_doble_inversor_24vcc.place(x=410, y=440, width=50, height=20)
lista_despl_rele_doble_inversor_24vcc.current(0)

lista_despl_rele_cuadru_inversor_12vcc = ttk.Combobox(values=["Caja","U"])
lista_despl_rele_cuadru_inversor_12vcc.place(x=410, y=470, width=50, height=20)
lista_despl_rele_cuadru_inversor_12vcc.current(0)

lista_despl_rele_cuadru_inversor_24vcc = ttk.Combobox(values=["Caja","U"])
lista_despl_rele_cuadru_inversor_24vcc.place(x=410, y=500, width=50, height=20)
lista_despl_rele_cuadru_inversor_24vcc.current(0)
    
lista_despl_rieldin = ttk.Combobox(values=["U","mts"])
lista_despl_rieldin.place(x=410, y=530, width=50, height=20)
lista_despl_rieldin.current(0)

lista_despl_terminales_sin_leng_0_75 = ttk.Combobox(values=["Caja","U"])
lista_despl_terminales_sin_leng_0_75.place(x=410, y=560, width=50, height=20)
lista_despl_terminales_sin_leng_0_75.current(0)

lista_despl_terminales_con_leng_2_5 = ttk.Combobox(values=["Caja","U"])
lista_despl_terminales_con_leng_2_5.place(x=410, y=590, width=50, height=20)
lista_despl_terminales_con_leng_2_5.current(0)

lista_despl_zocaro_rele_doble = ttk.Combobox(values=["Caja","U"])
lista_despl_zocaro_rele_doble.place(x=410, y=620, width=50, height=20)
lista_despl_zocaro_rele_doble.current(0)

lista_despl_zocaro_rele_cuadru = ttk.Combobox(values=["Caja","U"])
lista_despl_zocaro_rele_cuadru.place(x=410, y=650, width=50, height=20)
lista_despl_zocaro_rele_cuadru.current(0)

ventana_principal.mainloop()